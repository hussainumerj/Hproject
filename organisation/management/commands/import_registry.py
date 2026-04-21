from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User

from openpyxl import load_workbook

from organisation.models import Department, Team, Dependency


def slug_email(name: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", ".", name.strip().lower())
    cleaned = re.sub(r"\.+", ".", cleaned).strip(".")
    if not cleaned:
        cleaned = "user"
    return f"{cleaned}@example.com"


class Command(BaseCommand):
    help = "Import team registry data from the coursework Excel file."

    def add_arguments(self, parser):
        parser.add_argument(
            "xlsx_path",
            type=str,
            help="Path to the Excel file to import",
        )
        parser.add_argument(
            "--reset",
            action="store_true",
            help="Delete existing Team/Dependency/Department data before import",
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options["xlsx_path"])

        if not xlsx_path.exists():
            raise CommandError(f"File not found: {xlsx_path}")

        if options["reset"]:
            self.stdout.write(self.style.WARNING("Resetting existing organisation data..."))
            Dependency.objects.all().delete()
            Team.objects.all().delete()
            Department.objects.all().delete()

        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Excel file is empty.")

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = rows[1:]

        header_index = {name: i for i, name in enumerate(headers)}

        required_headers = [
            "Department",
            "Team Leader",
            "Department Head",
            "Team Name",
            "Project (codebase) (Github Repo)",
            "Downstream Dependencies",
            "Dependency Type",
            "Slack Channels",
            "Development Focus Areas",
        ]

        missing = [h for h in required_headers if h not in header_index]
        if missing:
            raise CommandError(f"Missing required columns: {', '.join(missing)}")

        created_users = 0
        created_departments = 0
        created_teams = 0
        created_dependencies = 0

        # First pass: create departments, users, teams
        team_dependency_buffer = []

        for row in data_rows:
            department_name = self._cell(row, header_index, "Department")
            team_leader_name = self._cell(row, header_index, "Team Leader")
            department_head_name = self._cell(row, header_index, "Department Head")
            team_name = self._cell(row, header_index, "Team Name")
            repo_url = self._cell(row, header_index, "Project (codebase) (Github Repo)")
            downstream_raw = self._cell(row, header_index, "Downstream Dependencies")
            dependency_type = self._cell(row, header_index, "Dependency Type")
            slack_channel = self._cell(row, header_index, "Slack Channels")
            focus_areas = self._cell(row, header_index, "Development Focus Areas")

            if not department_name or not team_name:
                continue

            leader_user = None
            if department_head_name:
                leader_user, was_created = User.objects.get_or_create(
                    username=self._username_from_name(department_head_name),
                    defaults={
                        "first_name": self._first_name(department_head_name),
                        "last_name": self._last_name(department_head_name),
                        "email": slug_email(department_head_name),
                    },
                )
                if was_created:
                    created_users += 1

            manager_user = None
            if team_leader_name:
                manager_user, was_created = User.objects.get_or_create(
                    username=self._username_from_name(team_leader_name),
                    defaults={
                        "first_name": self._first_name(team_leader_name),
                        "last_name": self._last_name(team_leader_name),
                        "email": slug_email(team_leader_name),
                    },
                )
                if was_created:
                    created_users += 1

            department, dept_created = Department.objects.get_or_create(
                name=department_name,
                defaults={
                    "description": f"Imported from registry for {department_name}.",
                    "specialisation": "",
                    "leader": leader_user,
                },
            )

            if dept_created:
                created_departments += 1
            else:
                # Keep first leader if already set, otherwise fill it
                if not department.leader and leader_user:
                    department.leader = leader_user
                    department.save()

            team, team_created = Team.objects.get_or_create(
                name=team_name,
                defaults={
                    "description": focus_areas or f"Imported team record for {team_name}.",
                    "purpose": dependency_type or "",
                    "department": department,
                    "manager": manager_user,
                    "slack_channel": slack_channel or "",
                    "email": manager_user.email if manager_user and manager_user.email else "",
                    "code_repository": repo_url or "",
                    "status": "active",
                },
            )

            if team_created:
                created_teams += 1
            else:
                # Update missing core fields if existing row is sparse
                changed = False
                if not team.department:
                    team.department = department
                    changed = True
                if not team.manager and manager_user:
                    team.manager = manager_user
                    changed = True
                if not team.description and focus_areas:
                    team.description = focus_areas
                    changed = True
                if not team.purpose and dependency_type:
                    team.purpose = dependency_type
                    changed = True
                if not team.slack_channel and slack_channel:
                    team.slack_channel = slack_channel
                    changed = True
                if not team.code_repository and repo_url:
                    team.code_repository = repo_url
                    changed = True
                if not team.email and manager_user and manager_user.email:
                    team.email = manager_user.email
                    changed = True
                if changed:
                    team.save()

            if manager_user:
                team.members.add(manager_user)

            if downstream_raw:
                team_dependency_buffer.append(
                    {
                        "team_name": team_name,
                        "downstream_names": [x.strip() for x in str(downstream_raw).split(",") if x and str(x).strip()],
                        "description": dependency_type or "",
                    }
                )

        # Second pass: create dependencies once all teams exist
        for item in team_dependency_buffer:
            try:
                upstream_team = Team.objects.get(name=item["team_name"])
            except Team.DoesNotExist:
                continue

            for downstream_name in item["downstream_names"]:
                try:
                    downstream_team = Team.objects.get(name=downstream_name)
                except Team.DoesNotExist:
                    self.stdout.write(
                        self.style.WARNING(
                            f"Skipped dependency: team '{downstream_name}' not found for '{upstream_team.name}'"
                        )
                    )
                    continue

                _, dep_created = Dependency.objects.get_or_create(
                    upstream_team=upstream_team,
                    downstream_team=downstream_team,
                    defaults={"description": item["description"]},
                )
                if dep_created:
                    created_dependencies += 1

        self.stdout.write(self.style.SUCCESS("Import complete."))
        self.stdout.write(f"Users created: {created_users}")
        self.stdout.write(f"Departments created: {created_departments}")
        self.stdout.write(f"Teams created: {created_teams}")
        self.stdout.write(f"Dependencies created: {created_dependencies}")

    def _cell(self, row, header_index, key):
        value = row[header_index[key]]
        if value is None:
            return ""
        return str(value).strip()

    def _username_from_name(self, full_name: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", full_name.strip().lower()) or "user"

    def _first_name(self, full_name: str) -> str:
        parts = full_name.strip().split()
        return parts[0] if parts else ""

    def _last_name(self, full_name: str) -> str:
        parts = full_name.strip().split()
        return " ".join(parts[1:]) if len(parts) > 1 else ""